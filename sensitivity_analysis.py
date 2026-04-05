"""
Parameter Sensitivity Analysis for WE Detect Spot Detection Pipeline.

Runs 19 configurations (1 baseline + 18 single-variable changes) and
saves results to an Excel file with conditional formatting.
"""

import sys
import importlib
from pathlib import Path
import cv2
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Path setup — must happen before importing device_drivers
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

# ---------------------------------------------------------------------------
# Import pipeline modules
# ---------------------------------------------------------------------------
import device_drivers.spot_analysis.config as cfg
import device_drivers.spot_analysis.detection as det_mod
import device_drivers.spot_analysis.inspection as insp_mod
from device_drivers.spot_analysis.detection import detect_spots, sort_and_label
from device_drivers.spot_analysis.inspection import inspect_spot_defects
from device_drivers.spot_analysis.visualization import draw_accept_reject_overlay

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
IMAGE_PATH = Path(
    r"C:\Users\Monster\Desktop\tez\writingPart\writing04.03"
    r"\ThesisImagesNotesSpotDetection\01_original.png"
)
OUTPUT_DIR = IMAGE_PATH.parent
EXCEL_PATH = OUTPUT_DIR / "parameter_sensitivity_results.xlsx"

# ---------------------------------------------------------------------------
# Default config snapshot (single source of truth)
# ---------------------------------------------------------------------------
DEFAULTS = {
    # Detection / geometric
    "DEFAULT_MIN_SPOT_AREA":    450,
    "DEFAULT_MAX_SPOT_AREA":    15000,
    "DEFAULT_MIN_CIRCULARITY":  0.45,
    "DEFAULT_MIN_SOLIDITY":     0.65,
    # Preprocessing
    "DEFAULT_BG_BLUR_K":        81,
    "DEFAULT_CLAHE_CLIP":       2.0,
    "DEFAULT_CLAHE_TILE":       (8, 8),
    "DEFAULT_THRESH_BLOCKSIZE": 35,
    "DEFAULT_THRESH_C":         2,
    "DEFAULT_OPEN_KERNEL":      2,
    "DEFAULT_CLOSE_KERNEL":     3,
    # Physical calibration
    "DEFAULT_MM_PER_PIXEL":     0.094,
    # Defect inspection
    "DEFAULT_MAD_K":            4.5,
    "DEFAULT_MAX_OUTLIER_FRAC": 0.16,
    "DEFAULT_DARK_Q":           10,
    "DEFAULT_BRIGHT_Q":         95,
    "DEFAULT_DEFECT_AREA_FRAC": 0.03,
    "DEFAULT_MIN_DEFECT_AREA_PX": 35,
}

# Variables that live on the detection module (imported via "from .config import")
DETECTION_VARS = {
    "DEFAULT_MIN_SPOT_AREA", "DEFAULT_MAX_SPOT_AREA",
    "DEFAULT_MIN_CIRCULARITY", "DEFAULT_MIN_SOLIDITY",
    "DEFAULT_BG_BLUR_K", "DEFAULT_CLAHE_CLIP", "DEFAULT_CLAHE_TILE",
    "DEFAULT_THRESH_BLOCKSIZE", "DEFAULT_THRESH_C",
    "DEFAULT_OPEN_KERNEL", "DEFAULT_CLOSE_KERNEL",
    "DEFAULT_MM_PER_PIXEL",
}

# Variables passed directly as kwargs to inspect_spot_defects()
INSPECTION_KWARGS = {
    "DEFAULT_MM_PER_PIXEL":     "mm_per_pixel",
    "DEFAULT_MAD_K":            "mad_k",
    "DEFAULT_MAX_OUTLIER_FRAC": "max_outlier_frac",
    "DEFAULT_DARK_Q":           "dark_q",
    "DEFAULT_BRIGHT_Q":         "bright_q",
    "DEFAULT_DEFECT_AREA_FRAC": "defect_area_frac",
    "DEFAULT_MIN_DEFECT_AREA_PX": "min_defect_area_px",
}

# ---------------------------------------------------------------------------
# Run definitions
# ---------------------------------------------------------------------------
RUNS = [
    # (run_id, group, variable, test_value)
    ("0",  "Baseline",              None,                        None),
    ("A1", "Geometric Filtering",   "DEFAULT_MIN_CIRCULARITY",   0.20),
    ("A2", "Geometric Filtering",   "DEFAULT_MIN_CIRCULARITY",   0.70),
    ("A3", "Geometric Filtering",   "DEFAULT_MIN_SOLIDITY",      0.40),
    ("A4", "Geometric Filtering",   "DEFAULT_MIN_SOLIDITY",      0.85),
    ("A5", "Geometric Filtering",   "DEFAULT_MIN_SPOT_AREA",     100),
    ("A6", "Geometric Filtering",   "DEFAULT_MIN_SPOT_AREA",     1000),
    ("B1", "Preprocessing",         "DEFAULT_THRESH_BLOCKSIZE",  15),
    ("B2", "Preprocessing",         "DEFAULT_THRESH_BLOCKSIZE",  61),
    ("B3", "Preprocessing",         "DEFAULT_THRESH_C",          0),
    ("B4", "Preprocessing",         "DEFAULT_THRESH_C",          5),
    ("B5", "Preprocessing",         "DEFAULT_CLAHE_CLIP",        1.0),
    ("B6", "Preprocessing",         "DEFAULT_CLAHE_CLIP",        4.0),
    ("C1", "Defect Inspection",     "DEFAULT_DEFECT_AREA_FRAC",  0.01),
    ("C2", "Defect Inspection",     "DEFAULT_DEFECT_AREA_FRAC",  0.10),
    ("C3", "Defect Inspection",     "DEFAULT_DARK_Q",            5),
    ("C4", "Defect Inspection",     "DEFAULT_DARK_Q",            20),
    ("C5", "Defect Inspection",     "DEFAULT_MM_PER_PIXEL",      0.047),
    ("C6", "Defect Inspection",     "DEFAULT_MM_PER_PIXEL",      0.188),
]


# ---------------------------------------------------------------------------
# Patch / restore helpers
# ---------------------------------------------------------------------------

def _apply_patch(var_name: str, value) -> None:
    """Set a config variable on both cfg and det_mod (where it's imported)."""
    setattr(cfg, var_name, value)
    if var_name in DETECTION_VARS:
        setattr(det_mod, var_name, value)


def _reset_all_to_defaults() -> None:
    """Restore every config variable to its default value."""
    for var, val in DEFAULTS.items():
        _apply_patch(var, val)


# ---------------------------------------------------------------------------
# Single-run executor
# ---------------------------------------------------------------------------

def run_once(image: np.ndarray, override_var: str | None, override_val) -> dict:
    """
    Run the full detection + inspection pipeline once.

    Detection module globals are patched before the call and restored via
    _reset_all_to_defaults() by the caller after.

    Inspection variables are passed directly as kwargs so no patching needed.
    """
    # Build inspection kwargs — start from defaults, override if needed
    insp_kwargs = {
        kwarg: DEFAULTS[cfg_var]
        for cfg_var, kwarg in INSPECTION_KWARGS.items()
    }
    if override_var and override_var in INSPECTION_KWARGS:
        insp_kwargs[INSPECTION_KWARGS[override_var]] = override_val

    # Run detection (uses patched module globals)
    spots, rejected_candidates, dbg = detect_spots(image)
    sort_and_label(spots)

    gray_norm = dbg["gray_norm"]

    # Run inspection
    accepted, rejected = [], []
    for s in spots:
        is_bad, metrics = inspect_spot_defects(gray_norm, s, **insp_kwargs)
        s["is_bad"] = is_bad
        s["metrics"] = metrics
        if is_bad:
            rejected.append(s)
        else:
            accepted.append(s)

    overlay = draw_accept_reject_overlay(image, spots)

    return {
        "detected":  len(spots),
        "accepted":  len(accepted),
        "rejected":  len(rejected),
        "overlay":   overlay,
    }


# ---------------------------------------------------------------------------
# Image filename builder
# ---------------------------------------------------------------------------

def _image_name(run_id: str, var: str | None, val) -> str:
    if run_id == "0":
        return "run_0_baseline.png"
    short = var.replace("DEFAULT_", "")
    val_str = str(val).replace(".", "_")
    return f"run_{run_id}_{short}_{val_str}.png"


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def _col(n: int) -> str:
    return get_column_letter(n)


def build_excel(rows: list[dict], run_configs: list[dict]) -> None:
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1 — Results                                                   #
    # ------------------------------------------------------------------ #
    ws = wb.active
    ws.title = "Results"

    header_fill   = PatternFill("solid", fgColor="1F3864")
    header_font   = Font(color="FFFFFF", bold=True)
    green_fill    = PatternFill("solid", fgColor="C6EFCE")
    red_fill      = PatternFill("solid", fgColor="FFC7CE")
    green_font    = Font(color="276221", bold=True)
    red_font      = Font(color="9C0006", bold=True)
    center        = Alignment(horizontal="center")
    thin          = Side(style="thin", color="AAAAAA")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Run ID", "Group", "Variable Changed",
        "Default Value", "Test Value",
        "Detected", "Accepted", "Rejected",
        "Δ Detected", "Δ Accepted",
    ]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = center
        cell.border = border

    col_delta_det = headers.index("Δ Detected") + 1
    col_delta_acc = headers.index("Δ Accepted") + 1

    for row_idx, r in enumerate(rows, start=2):
        values = [
            r["run_id"],
            r["group"],
            r["variable"] if r["variable"] else "—",
            r["default_val"] if r["default_val"] is not None else "—",
            r["test_val"]    if r["test_val"]    is not None else "—",
            r["detected"],
            r["accepted"],
            r["rejected"],
            r["delta_detected"],
            r["delta_accepted"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = center
            cell.border    = border

            # Conditional formatting on delta columns
            if col_idx in (col_delta_det, col_delta_acc):
                if isinstance(val, (int, float)) and val > 0:
                    cell.fill = green_fill
                    cell.font = green_font
                elif isinstance(val, (int, float)) and val < 0:
                    cell.fill = red_fill
                    cell.font = red_font

    # Column widths
    col_widths = [8, 22, 30, 16, 12, 11, 11, 11, 13, 13]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[_col(i)].width = w

    ws.freeze_panes = "A2"

    # ------------------------------------------------------------------ #
    # Sheet 2 — Run Config                                                #
    # ------------------------------------------------------------------ #
    ws2 = wb.create_sheet("Run Config")

    cfg_vars = list(DEFAULTS.keys())
    cfg_headers = ["Run ID", "Group", "Variable Changed"] + cfg_vars

    for col_idx, h in enumerate(cfg_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = border

    changed_fill = PatternFill("solid", fgColor="FFEB9C")
    changed_font = Font(color="9C5700", bold=True)

    for row_idx, rc in enumerate(run_configs, start=2):
        meta = [rc["run_id"], rc["group"], rc["variable"] if rc["variable"] else "—"]
        for col_idx, val in enumerate(meta, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = center
            cell.border    = border

        for v_idx, var in enumerate(cfg_vars, start=4):
            val  = rc["config"][var]
            cell = ws2.cell(row=row_idx, column=v_idx, value=str(val))
            cell.alignment = center
            cell.border    = border
            if rc["variable"] == var:
                cell.fill = changed_fill
                cell.font = changed_font

    ws2.freeze_panes = "D2"
    for col_idx in range(1, len(cfg_headers) + 1):
        ws2.column_dimensions[_col(col_idx)].width = 22

    wb.save(str(EXCEL_PATH))
    print(f"\nExcel saved → {EXCEL_PATH}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print(f"Loading image: {IMAGE_PATH}")
    image = cv2.imread(str(IMAGE_PATH))
    if image is None:
        raise FileNotFoundError(f"Cannot load image: {IMAGE_PATH}")
    print(f"Image loaded: {image.shape[1]}×{image.shape[0]} px\n")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    results     = []
    run_configs = []
    baseline    = None

    for run_id, group, var, val in RUNS:
        # --- reset everything to defaults first ---
        _reset_all_to_defaults()

        # --- build the config snapshot for this run ---
        config_snapshot = dict(DEFAULTS)

        if var is not None:
            config_snapshot[var] = val
            _apply_patch(var, val)

        # --- execute ---
        label = f"Run {run_id:>2} | {group:<22} | "
        if var:
            label += f"{var} = {val}"
        else:
            label += "BASELINE (all defaults)"
        print(label, end="  →  ", flush=True)

        result = run_once(image, var, val)

        # --- save annotated image ---
        img_name = _image_name(run_id, var, val)
        img_path = OUTPUT_DIR / img_name
        cv2.imwrite(str(img_path), result["overlay"])

        # --- delta vs baseline ---
        if run_id == "0":
            baseline = result

        delta_det = result["detected"] - baseline["detected"]
        delta_acc = result["accepted"] - baseline["accepted"]

        print(
            f"detected={result['detected']:3d}  "
            f"accepted={result['accepted']:3d}  "
            f"rejected={result['rejected']:3d}  "
            f"Δdet={delta_det:+d}  Δacc={delta_acc:+d}"
        )

        results.append({
            "run_id":          run_id,
            "group":           group,
            "variable":        var,
            "default_val":     DEFAULTS.get(var) if var else None,
            "test_val":        val,
            "detected":        result["detected"],
            "accepted":        result["accepted"],
            "rejected":        result["rejected"],
            "delta_detected":  delta_det,
            "delta_accepted":  delta_acc,
        })

        run_configs.append({
            "run_id":   run_id,
            "group":    group,
            "variable": var,
            "config":   config_snapshot,
        })

    # --- final reset ---
    _reset_all_to_defaults()

    # --- write Excel ---
    build_excel(results, run_configs)
    print("\nDone.")


if __name__ == "__main__":
    main()
