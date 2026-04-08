"""
Parameter Sensitivity Analysis — WE Detect Spot Pipeline.

Runs 19 configurations (1 baseline + 18 single-variable sweeps) on one image
and writes a 3-sheet Excel report plus one annotated PNG per run.

Output location:
  C:\\Users\\Monster\\Desktop\\tez\\writingPart\\writing04.03\\ThesisImagesNotesSpotDetection\\
"""

import sys
import cv2
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

IMAGE_PATH = Path(
    r"C:\Users\Monster\Desktop\tez\writingPart\writing04.03"
    r"\ThesisImagesNotesSpotDetection\01_original.png"
)
OUTPUT_DIR = IMAGE_PATH.parent
EXCEL_PATH = OUTPUT_DIR / "parameter_sensitivity_results.xlsx"

# ---------------------------------------------------------------------------
# Import pipeline modules AFTER path setup
# ---------------------------------------------------------------------------
import device_drivers.spot_analysis.config as cfg
import device_drivers.spot_analysis.detection as det_mod
from device_drivers.spot_analysis.detection import detect_spots, sort_and_label
from device_drivers.spot_analysis.inspection import inspect_spot_defects

# ---------------------------------------------------------------------------
# Default config snapshot (single source of truth for the script)
# ---------------------------------------------------------------------------
DEFAULTS = {
    # Geometric filters — live as module-level names in detection.py
    "DEFAULT_MIN_SPOT_AREA":      450,
    "DEFAULT_MAX_SPOT_AREA":      15000,
    "DEFAULT_MIN_CIRCULARITY":    0.45,
    "DEFAULT_MIN_SOLIDITY":       0.65,
    # Physical size filter — also live as module-level names in detection.py
    "DEFAULT_PLATE_WIDTH_MM":     50.0,
    "DEFAULT_MIN_SPOT_DIAMETER_MM": 1.5,
    # Preprocessing — module-level names in detection.py
    "DEFAULT_BG_BLUR_K":          81,
    "DEFAULT_CLAHE_CLIP":         2.0,
    "DEFAULT_CLAHE_TILE":         (8, 8),
    "DEFAULT_THRESH_BLOCKSIZE":   35,
    "DEFAULT_THRESH_C":           2,
    "DEFAULT_OPEN_KERNEL":        2,
    "DEFAULT_CLOSE_KERNEL":       3,
    # Defect inspection — these are DEFAULT ARGUMENTS in inspect_spot_defects,
    # so they must be passed as explicit kwargs; patching cfg won't help.
    "DEFAULT_ERODE_PX":           2,
    "DEFAULT_MAD_K":              4.5,
    "DEFAULT_MAX_OUTLIER_FRAC":   0.16,
    "DEFAULT_DARK_Q":             10,
    "DEFAULT_BRIGHT_Q":           95,
    "DEFAULT_DEFECT_AREA_FRAC":   0.03,
    "DEFAULT_MIN_DEFECT_AREA_PX": 35,
}

# Detection module vars: accessed by name inside function body → patching works.
DETECTION_VARS = {
    "DEFAULT_MIN_SPOT_AREA", "DEFAULT_MAX_SPOT_AREA",
    "DEFAULT_MIN_CIRCULARITY", "DEFAULT_MIN_SOLIDITY",
    "DEFAULT_PLATE_WIDTH_MM", "DEFAULT_MIN_SPOT_DIAMETER_MM",
    "DEFAULT_BG_BLUR_K", "DEFAULT_CLAHE_CLIP", "DEFAULT_CLAHE_TILE",
    "DEFAULT_THRESH_BLOCKSIZE", "DEFAULT_THRESH_C",
    "DEFAULT_OPEN_KERNEL", "DEFAULT_CLOSE_KERNEL",
}

# Inspection kwargs: inspect_spot_defects has these as default ARGUMENTS,
# so we MUST pass them explicitly — patching the module has no effect.
INSPECTION_KWARG_MAP = {
    "DEFAULT_ERODE_PX":           "erode_px",
    "DEFAULT_MAD_K":              "mad_k",
    "DEFAULT_MAX_OUTLIER_FRAC":   "max_outlier_frac",
    "DEFAULT_DARK_Q":             "dark_q",
    "DEFAULT_BRIGHT_Q":           "bright_q",
    "DEFAULT_DEFECT_AREA_FRAC":   "defect_area_frac",
    "DEFAULT_MIN_DEFECT_AREA_PX": "min_defect_area_px",
}

# ---------------------------------------------------------------------------
# Run definitions (19 total)
# ---------------------------------------------------------------------------
RUNS = [
    {"id": "0",  "group": "Baseline",           "variable": None,                        "default": None,  "test_value": None},
    {"id": "A1", "group": "Geometric Filtering", "variable": "DEFAULT_MIN_CIRCULARITY",   "default": 0.45,  "test_value": 0.20},
    {"id": "A2", "group": "Geometric Filtering", "variable": "DEFAULT_MIN_CIRCULARITY",   "default": 0.45,  "test_value": 0.70},
    {"id": "A3", "group": "Geometric Filtering", "variable": "DEFAULT_MIN_SOLIDITY",      "default": 0.65,  "test_value": 0.40},
    {"id": "A4", "group": "Geometric Filtering", "variable": "DEFAULT_MIN_SOLIDITY",      "default": 0.65,  "test_value": 0.85},
    {"id": "A5", "group": "Geometric Filtering", "variable": "DEFAULT_MIN_SPOT_AREA",     "default": 450,   "test_value": 100},
    {"id": "A6", "group": "Geometric Filtering", "variable": "DEFAULT_MIN_SPOT_AREA",     "default": 450,   "test_value": 1000},
    {"id": "B1", "group": "Preprocessing",       "variable": "DEFAULT_THRESH_BLOCKSIZE",  "default": 35,    "test_value": 15},
    {"id": "B2", "group": "Preprocessing",       "variable": "DEFAULT_THRESH_BLOCKSIZE",  "default": 35,    "test_value": 61},
    {"id": "B3", "group": "Preprocessing",       "variable": "DEFAULT_THRESH_C",          "default": 2,     "test_value": 0},
    {"id": "B4", "group": "Preprocessing",       "variable": "DEFAULT_THRESH_C",          "default": 2,     "test_value": 5},
    {"id": "B5", "group": "Preprocessing",       "variable": "DEFAULT_CLAHE_CLIP",        "default": 2.0,   "test_value": 1.0},
    {"id": "B6", "group": "Preprocessing",       "variable": "DEFAULT_CLAHE_CLIP",        "default": 2.0,   "test_value": 4.0},
    {"id": "C1", "group": "Defect Inspection",   "variable": "DEFAULT_DEFECT_AREA_FRAC",  "default": 0.03,  "test_value": 0.01},
    {"id": "C2", "group": "Defect Inspection",   "variable": "DEFAULT_DEFECT_AREA_FRAC",  "default": 0.03,  "test_value": 0.10},
    {"id": "C3", "group": "Defect Inspection",   "variable": "DEFAULT_DARK_Q",            "default": 10,    "test_value": 5},
    {"id": "C4", "group": "Defect Inspection",   "variable": "DEFAULT_DARK_Q",            "default": 10,    "test_value": 20},
    {"id": "C5", "group": "Defect Inspection",   "variable": "DEFAULT_ERODE_PX",          "default": 2,     "test_value": 0},
    {"id": "C6", "group": "Defect Inspection",   "variable": "DEFAULT_ERODE_PX",          "default": 2,     "test_value": 5},
]


# ---------------------------------------------------------------------------
# Patch / restore helpers
# ---------------------------------------------------------------------------

def _reset_all():
    """Restore every detection module variable to its default."""
    for var, val in DEFAULTS.items():
        if var in DETECTION_VARS:
            setattr(det_mod, var, val)
        setattr(cfg, var, val)


def _apply_override(var: str, val):
    if var in DETECTION_VARS:
        setattr(det_mod, var, val)
    setattr(cfg, var, val)


def _build_insp_kwargs(override_var, override_val) -> dict:
    """Build the full kwargs dict for inspect_spot_defects from current DEFAULTS,
    then apply the override if it's an inspection variable."""
    kwargs = {
        kwarg: DEFAULTS[cfg_var]
        for cfg_var, kwarg in INSPECTION_KWARG_MAP.items()
    }
    if override_var and override_var in INSPECTION_KWARG_MAP:
        kwargs[INSPECTION_KWARG_MAP[override_var]] = override_val
    return kwargs


# ---------------------------------------------------------------------------
# Single run executor
# ---------------------------------------------------------------------------

def run_once(image: np.ndarray, run: dict) -> dict:
    """Execute one full pipeline run and return summary + per-spot data."""
    var = run["variable"]
    val = run["test_value"]

    # 1. Reset everything
    _reset_all()

    # 2. Apply override to detection module (if applicable)
    if var is not None:
        _apply_override(var, val)
        actual = getattr(det_mod, var, None) if var in DETECTION_VARS else getattr(cfg, var, None)
        print(f"       override confirmed: {var} = {actual}")

    # 3. Build inspection kwargs (handles cached default-argument issue)
    insp_kwargs = _build_insp_kwargs(var, val)

    # 4. Run detection
    dbg: dict = {}
    spots, rejected_candidates, _ = detect_spots(image, debug=dbg)

    total_contours = dbg.get("total_contours", len(spots) + len(rejected_candidates))
    sort_and_label(spots)
    gray_norm = dbg["gray_norm"]

    # 5. Count geometric vs size rejects
    geom_rejected  = [r for r in rejected_candidates if not r.get("passed_geom", False)]
    size_rejected  = [r for r in rejected_candidates if r.get("passed_geom") and not r.get("passed_size", False)]
    passed_geom    = len(spots) + len(size_rejected)
    passed_size    = len(spots)

    # 6. Run defect inspection on each accepted candidate
    accepted_spots = []
    defect_rejected = []
    per_spot_rows = []

    for i, s in enumerate(spots):
        is_bad, metrics = inspect_spot_defects(gray_norm, s, **insp_kwargs)
        s["is_bad"]  = is_bad
        s["metrics"] = metrics
        cx, cy = s["center"]
        per_spot_rows.append({
            "spot_index":      i,
            "label":           s.get("label", ""),
            "center_x":        cx,
            "center_y":        cy,
            "radius_px":       round(s.get("radius_px", 0.0), 2),
            "diameter_mm":     round(s.get("diameter_mm", 0.0), 3),
            "passed_geometry": True,
            "passed_size":     True,
            "passed_defect":   not is_bad,
            "final_status":    "Accepted" if not is_bad else "Rejected",
            "rejection_reason": ", ".join(metrics.get("reason", [])) if is_bad else "",
        })
        if is_bad:
            defect_rejected.append(s)
        else:
            accepted_spots.append(s)

    # Add geom-rejected contours to per-spot detail
    for r in geom_rejected:
        per_spot_rows.append({
            "spot_index":      None,
            "label":           "—",
            "center_x":        None,
            "center_y":        None,
            "radius_px":       round(r.get("radius_px", 0.0), 2),
            "diameter_mm":     round(r.get("diameter_mm", 0.0), 3),
            "passed_geometry": False,
            "passed_size":     False,
            "passed_defect":   False,
            "final_status":    "Rejected",
            "rejection_reason": r.get("reason", ""),
        })

    # Add size-rejected contours to per-spot detail
    for r in size_rejected:
        per_spot_rows.append({
            "spot_index":      None,
            "label":           "—",
            "center_x":        None,
            "center_y":        None,
            "radius_px":       round(r.get("radius_px", 0.0), 2),
            "diameter_mm":     round(r.get("diameter_mm", 0.0), 3),
            "passed_geometry": True,
            "passed_size":     False,
            "passed_defect":   False,
            "final_status":    "Rejected",
            "rejection_reason": "too_small_physical",
        })

    # 7. Build annotated images
    overlay = image.copy()
    for s in accepted_spots:
        cx, cy = s["center"]
        r = int(s.get("radius_px", 10))
        cv2.circle(overlay, (cx, cy), r, (0, 200, 0), 2)
        cv2.putText(overlay, s.get("label", ""), (cx + r + 2, cy - 4),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 200, 0), 1)
    for s in defect_rejected:
        cx, cy = s["center"]
        r = int(s.get("radius_px", 10))
        cv2.circle(overlay, (cx, cy), r, (0, 0, 220), 2)
        cv2.putText(overlay, s.get("label", ""), (cx + r + 2, cy - 4),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 220), 1)
    for rc in geom_rejected + size_rejected:
        cnt = rc.get("contour")
        if cnt is not None:
            cv2.drawContours(overlay, [cnt], -1, (0, 200, 220), 1)

    # All-detected image: every spot that passed geometric + size filters in blue
    all_detected = image.copy()
    for s in spots:
        cx, cy = s["center"]
        r = int(s.get("radius_px", 10))
        cv2.circle(all_detected, (cx, cy), r, (255, 100, 0), 2)
        cv2.putText(all_detected, s.get("label", ""), (cx + r + 2, cy - 4),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 100, 0), 1)

    return {
        "total_contours":  total_contours,
        "detected_geom":   passed_geom,
        "passed_size":     passed_size,
        "accepted":        len(accepted_spots),
        "rejected":        len(defect_rejected) + len(size_rejected),
        "overlay":         overlay,
        "all_detected":    all_detected,
        "per_spot_rows":   per_spot_rows,
    }


# ---------------------------------------------------------------------------
# Image filename builder
# ---------------------------------------------------------------------------

def _img_name(run: dict) -> str:
    rid = run["id"]
    var = run["variable"]
    val = run["test_value"]
    if rid == "0":
        return "0_results.png"
    short = str(var).replace("DEFAULT_", "")
    val_s = str(val).replace(".", "_")
    return f"{rid}_{short}_{val_s}_results.png"


def _img_name_all(run: dict) -> str:
    rid = run["id"]
    var = run["variable"]
    val = run["test_value"]
    if rid == "0":
        return "0_all_detected.png"
    short = str(var).replace("DEFAULT_", "")
    val_s = str(val).replace(".", "_")
    return f"{rid}_{short}_{val_s}_all_detected.png"


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def _col(n: int) -> str:
    return get_column_letter(n)


def _header_style(cell, text):
    cell.value     = text
    cell.font      = Font(bold=True, color="FFFFFF")
    cell.fill      = PatternFill("solid", fgColor="1F3864")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _thin_border()


def _thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def _set_cell(ws, row, col, value, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = Alignment(horizontal="center")
    if border:
        c.border = _thin_border()
    return c


def build_excel(summary_rows: list, per_spot_rows_all: list, run_configs: list):
    wb = openpyxl.Workbook()
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    green_font = Font(color="276221", bold=True)
    red_fill   = PatternFill("solid", fgColor="FFC7CE")
    red_font   = Font(color="9C0006", bold=True)

    # ------------------------------------------------------------------ #
    # Sheet 1 — Summary                                                   #
    # ------------------------------------------------------------------ #
    ws1 = wb.active
    ws1.title = "Summary"

    s1_headers = [
        "Run", "Group", "Variable", "Default Value", "Test Value",
        "Total Contours", "Detected (geom)", "Passed Size",
        "Accepted", "Rejected", "Δ Detected", "Δ Accepted", "Notes",
    ]
    for ci, h in enumerate(s1_headers, 1):
        _header_style(ws1.cell(row=1, column=ci), h)

    delta_det_col = s1_headers.index("Δ Detected") + 1
    delta_acc_col = s1_headers.index("Δ Accepted") + 1

    for ri, row in enumerate(summary_rows, 2):
        vals = [
            row["run_id"],
            row["group"],
            row["variable"] or "—",
            row["default_val"] if row["default_val"] is not None else "—",
            row["test_val"]    if row["test_val"]    is not None else "—",
            row["total_contours"],
            row["detected_geom"],
            row["passed_size"],
            row["accepted"],
            row["rejected"],
            row["delta_detected"],
            row["delta_accepted"],
            "",
        ]
        for ci, v in enumerate(vals, 1):
            c = _set_cell(ws1, ri, ci, v)
            if ci == delta_det_col and isinstance(v, (int, float)):
                if v > 0:   c.fill, c.font = green_fill, green_font
                elif v < 0: c.fill, c.font = red_fill,   red_font
            if ci == delta_acc_col and isinstance(v, (int, float)):
                if v > 0:   c.fill, c.font = green_fill, green_font
                elif v < 0: c.fill, c.font = red_fill,   red_font

    col_widths_s1 = [6, 22, 30, 14, 12, 16, 17, 14, 11, 11, 13, 13, 14]
    for i, w in enumerate(col_widths_s1, 1):
        ws1.column_dimensions[_col(i)].width = w
    ws1.freeze_panes = "A2"

    # ------------------------------------------------------------------ #
    # Sheet 2 — Per-Spot Detail                                           #
    # ------------------------------------------------------------------ #
    ws2 = wb.create_sheet("Per-Spot Detail")

    s2_headers = [
        "Run", "Spot Index", "Spot Label", "Center X (px)", "Center Y (px)",
        "Radius (px)", "Diameter (mm)",
        "Passed Geometry", "Passed Size", "Passed Defect",
        "Final Status", "Rejection Reason",
    ]
    for ci, h in enumerate(s2_headers, 1):
        _header_style(ws2.cell(row=1, column=ci), h)

    r2 = 2
    for entry in per_spot_rows_all:
        run_id = entry["run_id"]
        for sp in entry["rows"]:
            vals = [
                run_id,
                sp["spot_index"] if sp["spot_index"] is not None else "—",
                sp.get("label", "—"),
                sp["center_x"]   if sp["center_x"]   is not None else "—",
                sp["center_y"]   if sp["center_y"]   is not None else "—",
                sp["radius_px"],
                sp["diameter_mm"],
                "Yes" if sp["passed_geometry"] else "No",
                "Yes" if sp["passed_size"]     else "No",
                "Yes" if sp["passed_defect"]   else "No",
                sp["final_status"],
                sp["rejection_reason"],
            ]
            for ci, v in enumerate(vals, 1):
                c = _set_cell(ws2, r2, ci, v)
                if ci == 11:  # Final Status
                    if v == "Accepted":
                        c.fill = green_fill; c.font = Font(color="276221", bold=True)
                    else:
                        c.fill = red_fill;   c.font = Font(color="9C0006", bold=True)
            r2 += 1

    col_widths_s2 = [6, 12, 12, 14, 14, 12, 14, 17, 13, 15, 14, 30]
    for i, w in enumerate(col_widths_s2, 1):
        ws2.column_dimensions[_col(i)].width = w
    ws2.freeze_panes = "A2"

    # ------------------------------------------------------------------ #
    # Sheet 3 — Run Config                                                #
    # ------------------------------------------------------------------ #
    ws3 = wb.create_sheet("Run Config")

    cfg_vars    = list(DEFAULTS.keys())
    s3_headers  = ["Run", "Group", "Variable Changed"] + cfg_vars
    for ci, h in enumerate(s3_headers, 1):
        _header_style(ws3.cell(row=1, column=ci), h)

    changed_fill = PatternFill("solid", fgColor="FFEB9C")
    changed_font = Font(color="9C5700", bold=True)

    for ri, rc in enumerate(run_configs, 2):
        meta = [rc["run_id"], rc["group"], rc["variable"] or "—"]
        for ci, v in enumerate(meta, 1):
            _set_cell(ws3, ri, ci, v)
        for vi, var in enumerate(cfg_vars, 4):
            v = rc["config"][var]
            c = _set_cell(ws3, ri, vi, str(v))
            if rc["variable"] == var:
                c.fill = changed_fill
                c.font = changed_font

    for i in range(1, len(s3_headers) + 1):
        ws3.column_dimensions[_col(i)].width = 22
    ws3.freeze_panes = "D2"

    wb.save(str(EXCEL_PATH))
    print(f"\nExcel saved → {EXCEL_PATH}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f"Loading image: {IMAGE_PATH}")
    image = cv2.imread(str(IMAGE_PATH))
    if image is None:
        raise FileNotFoundError(f"Cannot load: {IMAGE_PATH}")
    print(f"Image loaded: {image.shape[1]}×{image.shape[0]} px\n")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    summary_rows    = []
    per_spot_all    = []
    run_configs     = []
    baseline        = None

    for run in RUNS:
        rid = run["id"]
        var = run["variable"]
        val = run["test_value"]

        label = f"Run {rid:>2} | {run['group']:<22} | "
        label += f"{var} = {val}" if var else "BASELINE"
        print(label, end="\n", flush=True)

        result = run_once(image, run)

        # Save annotated images
        img_path = OUTPUT_DIR / _img_name(run)
        cv2.imwrite(str(img_path), result["overlay"])
        all_path = OUTPUT_DIR / _img_name_all(run)
        cv2.imwrite(str(all_path), result["all_detected"])

        if rid == "0":
            baseline = result

        delta_det = result["detected_geom"] - baseline["detected_geom"]
        delta_acc = result["accepted"]      - baseline["accepted"]

        print(
            f"       total={result['total_contours']}  "
            f"geom={result['detected_geom']}  "
            f"size={result['passed_size']}  "
            f"accepted={result['accepted']}  "
            f"rejected={result['rejected']}  "
            f"Δgeom={delta_det:+d}  Δacc={delta_acc:+d}"
        )

        # Build config snapshot for Run Config sheet
        config_snap = dict(DEFAULTS)
        if var:
            config_snap[var] = val

        summary_rows.append({
            "run_id":          rid,
            "group":           run["group"],
            "variable":        var,
            "default_val":     run["default"],
            "test_val":        val,
            "total_contours":  result["total_contours"],
            "detected_geom":   result["detected_geom"],
            "passed_size":     result["passed_size"],
            "accepted":        result["accepted"],
            "rejected":        result["rejected"],
            "delta_detected":  delta_det,
            "delta_accepted":  delta_acc,
        })

        per_spot_all.append({
            "run_id": rid,
            "rows":   result["per_spot_rows"],
        })

        run_configs.append({
            "run_id":   rid,
            "group":    run["group"],
            "variable": var,
            "config":   config_snap,
        })

    _reset_all()
    build_excel(summary_rows, per_spot_all, run_configs)
    print("\nDone.")


if __name__ == "__main__":
    main()
