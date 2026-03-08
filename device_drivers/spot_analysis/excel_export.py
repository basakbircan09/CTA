# excel_export.py

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import Dict, Any


# ---------------------------------------------------------------------------
# Colour fills
# ---------------------------------------------------------------------------
_FILL_ACCEPTED      = PatternFill("solid", fgColor="DDEBF7")   # light blue
_FILL_REJECTED      = PatternFill("solid", fgColor="FDE9E7")   # light red
_FILL_HDR_BLUE      = PatternFill("solid", fgColor="D9EAF7")
_FILL_HDR_ORANGE    = PatternFill("solid", fgColor="FCE4D6")
_FILL_HDR_GREEN     = PatternFill("solid", fgColor="E2F0D9")


def _write_header_row(ws, headers: list, fill: PatternFill) -> None:
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 32)


# Columns shared by All_Spots and Rejected_Spots sheets
_SPOT_HEADERS = [
    "Label",
    "Status",
    "Center_X",
    "Center_Y",
    "Area",
    "Circularity",
    "Solidity",
    "Median_Brightness",
    "MAD",
    "Outlier_Fraction",
    "Nonuniform_Flag",
    "Dark_Threshold",
    "Bright_Threshold",
    "Min_Defect_Area_px",
    "Reject_Reasons",
    "Warning",
    "Pixel_Count_Inside_Check",
]


def _spot_row(s: dict) -> list:
    metrics = s.get("metrics", {})
    return [
        s.get("label", ""),
        "Rejected" if s.get("is_bad", False) else "Accepted",
        s["center"][0],
        s["center"][1],
        s.get("area", ""),
        s.get("circularity", ""),
        s.get("solidity", ""),
        metrics.get("median", ""),
        metrics.get("mad", ""),
        metrics.get("outlier_frac", ""),
        metrics.get("nonuniform_flag", ""),
        metrics.get("t_dark", ""),
        metrics.get("t_bright", ""),
        metrics.get("min_defect_area_px", ""),
        ", ".join(metrics.get("reason", [])),
        metrics.get("warning", ""),
        metrics.get("n", ""),
    ]


def _fill_spot_sheet(ws, spots: list, header_fill: PatternFill) -> None:
    _write_header_row(ws, _SPOT_HEADERS, header_fill)
    for r, s in enumerate(spots, start=2):
        row_data = _spot_row(s)
        fill = _FILL_REJECTED if s.get("is_bad", False) else _FILL_ACCEPTED
        for c, value in enumerate(row_data, start=1):
            cell = ws.cell(r, c, value)
            cell.fill = fill
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_results_to_excel(path: str, result: Dict[str, Any]) -> None:
    """Write spot analysis results to a styled 4-sheet .xlsx workbook.

    Sheets
    ------
    Summary                       – counts + label lists
    All_Spots                     – every accepted spot with full metrics
    Rejected_Spots                – only defective spots
    Detection_Rejected_Candidates – contours filtered during detection
    """
    wb = Workbook()

    # ---- Summary ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Spot Detection Summary"
    ws["A1"].font = Font(bold=True, size=14)

    summary_rows = [
        ("Detected spots",                        len(result["all_spots"])),
        ("Accepted spots",                        len(result["accepted_spots"])),
        ("Rejected spots",                        len(result["rejected_spots"])),
        ("Detection-stage filtered candidates",   len(result["rejected_candidates"])),
    ]
    for i, (label, value) in enumerate(summary_rows, start=3):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value

    # optional label/missing rows
    if result.get("accepted_labels"):
        ws.append(["Accepted Labels", ", ".join(result["accepted_labels"])])
    if result.get("rejected_labels"):
        ws.append(["Rejected Labels", ", ".join(result["rejected_labels"])])
    if result.get("missing_spots"):
        ws.append(["Missing Spots",   ", ".join(result["missing_spots"])])

    ws.freeze_panes = "A3"

    # ---- All_Spots ----
    ws2 = wb.create_sheet("All_Spots")
    _fill_spot_sheet(ws2, result["all_spots"], _FILL_HDR_BLUE)

    # ---- Rejected_Spots ----
    ws3 = wb.create_sheet("Rejected_Spots")
    rejected_only = [s for s in result["all_spots"] if s.get("is_bad", False)]
    _fill_spot_sheet(ws3, rejected_only, _FILL_HDR_ORANGE)

    # ---- Detection_Rejected_Candidates ----
    ws4 = wb.create_sheet("Detection_Rejected_Candidates")
    cand_headers = ["Reason", "Area", "Circularity", "Solidity"]
    _write_header_row(ws4, cand_headers, _FILL_HDR_GREEN)
    for r, s in enumerate(result["rejected_candidates"], start=2):
        ws4.cell(r, 1, s.get("reason", ""))
        ws4.cell(r, 2, s.get("area", ""))
        ws4.cell(r, 3, s.get("circularity", ""))
        ws4.cell(r, 4, s.get("solidity", ""))
    ws4.freeze_panes = "A2"

    for sheet in [ws, ws2, ws3, ws4]:
        _autosize(sheet)

    wb.save(path)
