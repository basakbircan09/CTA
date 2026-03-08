import cv2
import numpy as np
from pathlib import Path
from typing import Dict, Any, List, Tuple
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ================= DEFAULT SETTINGS =================
# Detection filters
DEFAULT_MIN_SPOT_AREA = 450
DEFAULT_MAX_SPOT_AREA = 15000
DEFAULT_MIN_CIRCULARITY = 0.45
DEFAULT_MIN_SOLIDITY = 0.65

# Preprocessing
DEFAULT_BG_BLUR_K = 81       # odd; larger -> stronger illumination correction
DEFAULT_CLAHE_CLIP = 2.0
DEFAULT_CLAHE_TILE = (8, 8)

# Thresholding
DEFAULT_THRESH_BLOCKSIZE = 35   # must be odd
DEFAULT_THRESH_C = 2

# Morphology
DEFAULT_OPEN_KERNEL = 2
DEFAULT_CLOSE_KERNEL = 3

# Defect inspection
DEFAULT_ERODE_PX = 2
DEFAULT_MAD_K = 4.5
DEFAULT_MAX_OUTLIER_FRAC = 0.16

DEFAULT_DARK_Q = 10
DEFAULT_BRIGHT_Q = 95
DEFAULT_DEFECT_AREA_FRAC = 0.03
DEFAULT_MIN_DEFECT_AREA_PX = 35
# ====================================================


def save_step(save_dir: Path, name: str, img) -> None:
    if save_dir is None:
        return
    save_dir.mkdir(parents=True, exist_ok=True)
    cv2.imwrite(str(save_dir / name), img)


def preprocess_for_detection(
    bgr: np.ndarray,
    bg_blur_k: int = DEFAULT_BG_BLUR_K,
    clahe_clip: float = DEFAULT_CLAHE_CLIP,
    clahe_tile: Tuple[int, int] = DEFAULT_CLAHE_TILE,
    debug: Dict[str, Any] = None
) -> np.ndarray:
    gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)

    k = bg_blur_k if bg_blur_k % 2 == 1 else bg_blur_k + 1
    bg = cv2.GaussianBlur(gray, (k, k), 0)
    bg = np.clip(bg, 1, 255).astype(np.uint8)

    # illumination normalization
    norm = cv2.divide(gray, bg, scale=255)

    # contrast normalization
    clahe = cv2.createCLAHE(clipLimit=clahe_clip, tileGridSize=clahe_tile)
    norm = clahe.apply(norm)

    if debug is not None:
        debug["gray_raw"] = gray
        debug["bg"] = bg
        debug["gray_norm"] = norm

    return norm


def detect_spots(
    plate_img: np.ndarray,
    min_area: int = DEFAULT_MIN_SPOT_AREA,
    max_area: int = DEFAULT_MAX_SPOT_AREA,
    min_circularity: float = DEFAULT_MIN_CIRCULARITY,
    min_solidity: float = DEFAULT_MIN_SOLIDITY,
    thresh_blocksize: int = DEFAULT_THRESH_BLOCKSIZE,
    thresh_c: int = DEFAULT_THRESH_C,
    open_kernel: int = DEFAULT_OPEN_KERNEL,
    close_kernel: int = DEFAULT_CLOSE_KERNEL,
    debug: Dict[str, Any] = None
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    pdbg = {}
    norm = preprocess_for_detection(plate_img, debug=pdbg)

    blur = cv2.GaussianBlur(norm, (5, 5), 0)

    blocksize = thresh_blocksize if thresh_blocksize % 2 == 1 else thresh_blocksize + 1
    thresh = cv2.adaptiveThreshold(
        blur, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV,
        blocksize, thresh_c
    )

    if open_kernel > 0:
        opened = cv2.morphologyEx(
            thresh,
            cv2.MORPH_OPEN,
            np.ones((open_kernel, open_kernel), np.uint8)
        )
    else:
        opened = thresh.copy()

    if close_kernel > 0:
        closed = cv2.morphologyEx(
            opened,
            cv2.MORPH_CLOSE,
            np.ones((close_kernel, close_kernel), np.uint8)
        )
    else:
        closed = opened.copy()

    if debug is not None:
        debug.update(pdbg)
        debug["blur"] = blur
        debug["thresh_bw"] = thresh
        debug["opened"] = opened
        debug["closed"] = closed

    contours, _ = cv2.findContours(closed, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    spots = []
    rejected_candidates = []

    for c in contours:
        area = cv2.contourArea(c)
        peri = cv2.arcLength(c, True)
        circ = 0.0 if peri <= 1e-6 else 4 * np.pi * area / (peri ** 2)

        hull = cv2.convexHull(c)
        hull_area = cv2.contourArea(hull)
        solidity = 0.0 if hull_area <= 1e-6 else area / hull_area

        reason = None

        if not (min_area <= area <= max_area):
            reason = "area"
        elif peri <= 1e-6:
            reason = "perimeter"
        elif circ < min_circularity:
            reason = "circularity"
        elif hull_area <= 1e-6:
            reason = "hull_area"
        elif solidity < min_solidity:
            reason = "solidity"

        if reason is not None:
            rejected_candidates.append({
                "contour": c,
                "reason": reason,
                "area": float(area),
                "circularity": float(circ),
                "solidity": float(solidity)
            })
            continue

        M = cv2.moments(c)
        if M["m00"] == 0:
            rejected_candidates.append({
                "contour": c,
                "reason": "zero_moment",
                "area": float(area),
                "circularity": float(circ),
                "solidity": float(solidity)
            })
            continue

        cx = int(M["m10"] / M["m00"])
        cy = int(M["m01"] / M["m00"])

        spots.append({
            "contour": c,
            "center": (cx, cy),
            "area": float(area),
            "circularity": float(circ),
            "solidity": float(solidity),
        })

    return spots, rejected_candidates


def _spot_mask(shape_hw, contour) -> np.ndarray:
    m = np.zeros(shape_hw, dtype=np.uint8)
    cv2.drawContours(m, [contour], -1, 255, thickness=-1)
    return m


def _erode_mask(mask: np.ndarray, erode_px: int) -> np.ndarray:
    if erode_px <= 0:
        return mask
    k = 2 * erode_px + 1
    ker = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (k, k))
    return cv2.erode(mask, ker, iterations=1)


def inspect_spot_defects(
    gray_norm: np.ndarray,
    spot: Dict[str, Any],
    erode_px: int = DEFAULT_ERODE_PX,
    mad_k: float = DEFAULT_MAD_K,
    max_outlier_frac: float = DEFAULT_MAX_OUTLIER_FRAC,
    dark_q: float = DEFAULT_DARK_Q,
    bright_q: float = DEFAULT_BRIGHT_Q,
    defect_area_frac: float = DEFAULT_DEFECT_AREA_FRAC,
    min_defect_area_px: int = DEFAULT_MIN_DEFECT_AREA_PX,
) -> Tuple[bool, Dict[str, Any]]:
    spot_mask = _spot_mask(gray_norm.shape, spot["contour"])
    inner = _erode_mask(spot_mask, erode_px)

    vals = gray_norm[inner == 255]
    if vals.size < 80:
        return False, {
            "reason": [],
            "n": int(vals.size),
            "warning": "too_few_pixels_for_reliable_defect_check"
        }

    med = float(np.median(vals))
    mad = float(np.median(np.abs(vals - med))) + 1e-6
    z = np.abs(vals - med) / (1.4826 * mad)
    outlier_frac = float(np.mean(z > mad_k))

    reasons = []
    metrics = {
        "median": med,
        "mad": mad,
        "outlier_frac": outlier_frac,
        "n": int(vals.size)
    }

    # informational only
    nonuniform_flag = outlier_frac > max_outlier_frac

    t_dark = float(np.percentile(vals, dark_q))
    t_bright = float(np.percentile(vals, bright_q))
    metrics["t_dark"] = t_dark
    metrics["t_bright"] = t_bright

    dark_bin = np.zeros_like(gray_norm, dtype=np.uint8)
    bright_bin = np.zeros_like(gray_norm, dtype=np.uint8)
    dark_bin[(inner == 255) & (gray_norm <= t_dark)] = 255
    bright_bin[(inner == 255) & (gray_norm >= t_bright)] = 255

    ker = np.ones((3, 3), np.uint8)
    dark_bin = cv2.morphologyEx(dark_bin, cv2.MORPH_OPEN, ker)
    bright_bin = cv2.morphologyEx(bright_bin, cv2.MORPH_OPEN, ker)

    inner_boundary = cv2.morphologyEx(inner, cv2.MORPH_GRADIENT, ker)

    spot_area_inner = int(np.sum(inner == 255))
    min_area = max(min_defect_area_px, int(defect_area_frac * spot_area_inner))
    metrics["min_defect_area_px"] = int(min_area)

    def has_valid_component(bin_img: np.ndarray) -> bool:
        num, labels, stats, _ = cv2.connectedComponentsWithStats(bin_img, connectivity=8)
        if num <= 1:
            return False
        for lab in range(1, num):
            area = int(stats[lab, cv2.CC_STAT_AREA])
            if area < min_area:
                continue
            comp = (labels == lab).astype(np.uint8) * 255
            touches = np.any((comp == 255) & (inner_boundary == 255))
            if not touches:
                return True
        return False

    if has_valid_component(dark_bin):
        reasons.append("dark_defect_component")
    if has_valid_component(bright_bin):
        reasons.append("bright_defect_component")

    metrics["nonuniform_flag"] = nonuniform_flag
    metrics["reason"] = reasons
    return (len(reasons) > 0), metrics


def sort_and_label(spots: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not spots:
        return []

    spots = sorted(spots, key=lambda s: s["center"][1])

    rows = []
    y_vals = [s["center"][1] for s in spots]
    row_thresh = np.median(np.diff(y_vals)) * 0.6 if len(y_vals) > 2 else 40

    for s in spots:
        for row in rows:
            if abs(s["center"][1] - row[0]["center"][1]) < row_thresh:
                row.append(s)
                break
        else:
            rows.append([s])

    labeled = []
    for r, row in enumerate(rows):
        row = sorted(row, key=lambda s: s["center"][0])
        for c, s in enumerate(row):
            s["label"] = f"{chr(65 + r)}{c + 1}"
            labeled.append(s)

    return labeled


def draw_accept_reject_overlay(image: np.ndarray, spots: List[Dict[str, Any]]) -> np.ndarray:
    out = image.copy()
    for s in spots:
        bad = s.get("is_bad", False)
        color = (0, 0, 255) if bad else (255, 0, 0)  # red=rejected, blue=accepted
        cv2.drawContours(out, [s["contour"]], -1, color, 2)
        cx, cy = s["center"]
        cv2.circle(out, (cx, cy), 3, (0, 0, 255), -1)
        cv2.putText(
            out, s["label"], (cx + 5, cy - 5),
            cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 1
        )
    return out


def draw_rejected_candidates_overlay(
    image: np.ndarray,
    rejected_candidates: List[Dict[str, Any]]
) -> np.ndarray:
    out = image.copy()
    for s in rejected_candidates:
        cv2.drawContours(out, [s["contour"]], -1, (0, 255, 255), 2)  # yellow
    return out


def print_detection_reject_summary(rejected_candidates: List[Dict[str, Any]]) -> None:
    counts = Counter([s["reason"] for s in rejected_candidates])

    print("\nDetection-stage rejected candidates:")
    if not counts:
        print("  None")
        return

    for reason, count in sorted(counts.items()):
        print(f"  {reason}: {count}")


def print_classification_summary(spots: List[Dict[str, Any]]) -> None:
    reason_counter = Counter()
    warning_counter = Counter()

    for s in spots:
        metrics = s.get("metrics", {})
        reasons = metrics.get("reason", [])
        warning = metrics.get("warning", None)

        for r in reasons:
            reason_counter[r] += 1
        if warning:
            warning_counter[warning] += 1

    print("\nClassification-stage reject reasons:")
    if not reason_counter:
        print("  None")
    else:
        for reason, count in sorted(reason_counter.items()):
            print(f"  {reason}: {count}")

    print("\nClassification warnings:")
    if not warning_counter:
        print("  None")
    else:
        for warning, count in sorted(warning_counter.items()):
            print(f"  {warning}: {count}")


def autosize_worksheet(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 32)


def export_results_to_excel(
    out_xlsx_path: str,
    result: Dict[str, Any]
) -> None:
    wb = Workbook()

    accepted_fill = PatternFill("solid", fgColor="DDEBF7")
    rejected_fill = PatternFill("solid", fgColor="FDE9E7")
    header_blue_fill = PatternFill("solid", fgColor="D9EAF7")
    header_orange_fill = PatternFill("solid", fgColor="FCE4D6")
    header_green_fill = PatternFill("solid", fgColor="E2F0D9")

    # ---------------- Summary ----------------
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "Spot Detection Summary"
    ws["A1"].font = Font(bold=True, size=14)

    summary_rows = [
        ("Detected spots", len(result["all_spots"])),
        ("Accepted spots", len(result["accepted_spots"])),
        ("Rejected spots", len(result["rejected_spots"])),
        ("Detection-stage filtered candidates", len(result["rejected_candidates"])),
    ]

    for i, (label, value) in enumerate(summary_rows, start=3):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value

    ws.freeze_panes = "A3"

    # ---------------- All spots ----------------
    ws2 = wb.create_sheet("All_Spots")

    headers = [
        "Label",
        "Status",
        "Center_X",
        "Center_Y",
        "Area",
        "Circularity",
        "Solidity",
        "Median_Brightness",
        "MAD",
        "Outlier_Fraction",
        "Nonuniform_Flag",
        "Dark_Threshold",
        "Bright_Threshold",
        "Min_Defect_Area_px",
        "Reject_Reasons",
        "Warning",
        "Pixel_Count_Inside_Check"
    ]

    for col, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_blue_fill
        cell.alignment = Alignment(horizontal="center")

    for r, s in enumerate(result["all_spots"], start=2):
        metrics = s.get("metrics", {})
        reasons = metrics.get("reason", [])
        warning = metrics.get("warning", "")
        n_pixels = metrics.get("n", "")

        status = "Rejected" if s.get("is_bad", False) else "Accepted"

        row_values = [
            s.get("label", ""),
            status,
            s["center"][0],
            s["center"][1],
            s.get("area", ""),
            s.get("circularity", ""),
            s.get("solidity", ""),
            metrics.get("median", ""),
            metrics.get("mad", ""),
            metrics.get("outlier_frac", ""),
            metrics.get("nonuniform_flag", ""),
            metrics.get("t_dark", ""),
            metrics.get("t_bright", ""),
            metrics.get("min_defect_area_px", ""),
            ", ".join(reasons),
            warning,
            n_pixels
        ]

        for c, value in enumerate(row_values, start=1):
            ws2.cell(r, c, value)

        fill = rejected_fill if status == "Rejected" else accepted_fill
        for c in range(1, len(headers) + 1):
            ws2.cell(r, c).fill = fill

    ws2.freeze_panes = "A2"

    # ---------------- Rejected spots only ----------------
    ws3 = wb.create_sheet("Rejected_Spots")

    rejected_headers = headers
    for col, h in enumerate(rejected_headers, start=1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_orange_fill
        cell.alignment = Alignment(horizontal="center")

    rejected_only = [s for s in result["all_spots"] if s.get("is_bad", False)]
    for r, s in enumerate(rejected_only, start=2):
        metrics = s.get("metrics", {})
        reasons = metrics.get("reason", [])
        warning = metrics.get("warning", "")
        n_pixels = metrics.get("n", "")

        row_values = [
            s.get("label", ""),
            "Rejected",
            s["center"][0],
            s["center"][1],
            s.get("area", ""),
            s.get("circularity", ""),
            s.get("solidity", ""),
            metrics.get("median", ""),
            metrics.get("mad", ""),
            metrics.get("outlier_frac", ""),
            metrics.get("nonuniform_flag", ""),
            metrics.get("t_dark", ""),
            metrics.get("t_bright", ""),
            metrics.get("min_defect_area_px", ""),
            ", ".join(reasons),
            warning,
            n_pixels
        ]

        for c, value in enumerate(row_values, start=1):
            ws3.cell(r, c, value)
            ws3.cell(r, c).fill = rejected_fill

    ws3.freeze_panes = "A2"

    # ---------------- Detection rejected candidates ----------------
    ws4 = wb.create_sheet("Detection_Rejected_Candidates")

    headers4 = [
        "Reason",
        "Area",
        "Circularity",
        "Solidity"
    ]

    for col, h in enumerate(headers4, start=1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_green_fill
        cell.alignment = Alignment(horizontal="center")

    for r, s in enumerate(result["rejected_candidates"], start=2):
        ws4.cell(r, 1, s.get("reason", ""))
        ws4.cell(r, 2, s.get("area", ""))
        ws4.cell(r, 3, s.get("circularity", ""))
        ws4.cell(r, 4, s.get("solidity", ""))

    ws4.freeze_panes = "A2"

    for sheet in [ws, ws2, ws3, ws4]:
        autosize_worksheet(sheet)

    wb.save(out_xlsx_path)


def analyze_spots_on_full_plate_image(
    image_path: str,
    save_dir: str = None,
) -> Dict[str, Any]:
    save_path = Path(save_dir) if save_dir else None

    img = cv2.imread(str(image_path))
    if img is None:
        return {"ok": False, "error": f"Image not found: {image_path}"}

    save_step(save_path, "01_original.png", img)

    det_dbg = {}
    spots, rejected_candidates = detect_spots(img, debug=det_dbg)
    spots = sort_and_label(spots)

    save_step(save_path, "02_gray_raw.png", det_dbg["gray_raw"])
    save_step(save_path, "03_bg.png", det_dbg["bg"])
    save_step(save_path, "04_gray_norm.png", det_dbg["gray_norm"])
    save_step(save_path, "05_thresh_bw.png", det_dbg["thresh_bw"])
    save_step(save_path, "06_opened.png", det_dbg["opened"])
    save_step(save_path, "07_closed.png", det_dbg["closed"])

    rejected_candidate_overlay = draw_rejected_candidates_overlay(img, rejected_candidates)
    save_step(save_path, "08_rejected_candidates_overlay.png", rejected_candidate_overlay)

    gray_norm = det_dbg["gray_norm"]
    accepted, rejected = [], []
    per_spot_metrics = {}

    for s in spots:
        is_bad, metrics = inspect_spot_defects(gray_norm, s)
        s["is_bad"] = is_bad
        s["metrics"] = metrics
        per_spot_metrics[s["label"]] = metrics

        if is_bad:
            rejected.append(s)
        else:
            accepted.append(s)

    overlay = draw_accept_reject_overlay(img, spots)
    save_step(save_path, "09_accept_reject_overlay.png", overlay)

    return {
        "ok": True,
        "all_spots": spots,
        "accepted_spots": accepted,
        "rejected_spots": rejected,
        "rejected_candidates": rejected_candidates,
        "per_spot_metrics": per_spot_metrics,
        "error": None
    }


def run_spot_analysis(
    image_path: str,
    output_dir: str = None,
    export_excel: bool = True,
    show_overlay: bool = False
) -> Dict[str, Any]:

    result = analyze_spots_on_full_plate_image(image_path, save_dir=output_dir)

    if not result["ok"]:
        return result

    print(f"Detected: {len(result['all_spots'])}")
    print(f"Accepted: {len(result['accepted_spots'])}")
    print(f"Rejected: {len(result['rejected_spots'])}")
    print(f"Detection-stage filtered candidates: {len(result['rejected_candidates'])}")

    print_detection_reject_summary(result["rejected_candidates"])
    print_classification_summary(result["all_spots"])

    # Export Excel if requested
    excel_path = None
    if export_excel and output_dir is not None:
        excel_path = str(Path(output_dir) / "spot_analysis_results.xlsx")
        export_results_to_excel(excel_path, result)

    if output_dir is not None:
        print(f"\nSaved debug images to: {output_dir}")
        print(f"Rejected-candidate overlay: {Path(output_dir) / '08_rejected_candidates_overlay.png'}")
        print(f"Accept/reject overlay: {Path(output_dir) / '09_accept_reject_overlay.png'}")

    if excel_path:
        print(f"Excel file: {excel_path}")

    if show_overlay and output_dir is not None:
        overlay = cv2.imread(str(Path(output_dir) / "09_accept_reject_overlay.png"))
        if overlay is not None:
            cv2.imshow("Accept (blue) / Reject (red)", overlay)
            cv2.waitKey(0)
            cv2.destroyAllWindows()

    return result

if __name__ == "__main__":

    image_path = input("Enter image path: ")
    output_dir = input("Enter output folder: ")

    run_spot_analysis(
        image_path=image_path,
        output_dir=output_dir,
        export_excel=True,
        show_overlay=True
    )